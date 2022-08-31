import datetime
import json
from os import path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.chart import BarChart, Reference

# 定义表标题
title = ['主机列表', '发行版本', '内核版本', 'selinux', '时间与时区', '系统参数优化', '磁盘空间使用', 'inode占用',
         '内存占用', 'swap占用', 'OPS节点', 'Node节点', 'Namespace']  # 定义表头
# 获取当前日期
d1 = datetime.datetime.now()
T = "{:%Y%m%d}".format(d1)

# 新建一个表格
workbook = openpyxl.Workbook(r'/home/gu/桌面/Python/venv/学习/venv/自动化.xlsx')
sheet = workbook.create_sheet('sheet')
sheet.append(title)
sheet.title = '黑屏'
workbook.save(r'/home/gu/桌面/Python/venv/学习/venv/自动化.xlsx')
# 从巡检报告中拿取所有的主机列表

# 定义主机列表
IP = []
# 每次运行前清除主机列表
IP.clear()
with open(r'xunjian/hosts/hosts', 'r+', encoding='utf8') as t1:  # 打开主机清单
    for ip in t1:  # 遍历主机清单
        ip1 = ip.replace('\n', "")  # 将主机清单中\n去除
        IP.append(ip1)  # 添加ip到主机清单列表

# 定义输出信息列表
std = []
# 将数据按ip依次写入表格
c = 0
for q1 in IP:  # 从主机清单遍历
    fn = path.join(r'xunjian/data/', T, q1)  # 定义数据路径
    with open(fn, 'r+') as f1:  # 依次打开路径下主机名称数据
        std.clear()
        for line1 in f1:  # 逐行读取数据
            with open(r'o.json', 'w+') as o:  # 每次以新文件形式打开一个新文件
                o.write(line1)  # 将读取的行写入新文件
            with open(r'o.json', 'r+') as o1:  # 打开刚刚写入的json文件
                k1 = json.load(o1)  # 将json格式反序列化为字符串
                if k1 is not None:
                    std1 = k1['stdout']  # 取出json中的对应数据
                    std.append(std1)  # 将数据添加到数据列表
        std.insert(0, IP[c])
        c += 1
        workbook1 = openpyxl.load_workbook(r'/home/gu/桌面/Python/venv/学习/venv/自动化.xlsx')
        sheet1 = workbook1['黑屏']
        sheet1.append(std)
        workbook1.save(r'/home/gu/桌面/Python/venv/学习/venv/自动化.xlsx')

# 制作表头
workbook3 = openpyxl.load_workbook(r'/home/gu/桌面/Python/venv/学习/venv/自动化.xlsx')
sheet3 = workbook3['黑屏']
sheet3.insert_rows(idx=1, amount=3)
sheet3.merge_cells(start_row=1, end_row=3, start_column=1, end_column=13)
sheet3['A1'] = ''
# 设置对齐样式
ali = Alignment(horizontal='center', vertical="center", wrap_text=True)
sheet3['A1'].alignment = ali
# 设置字体
font = Font(name='宋体', size='36', bold=True)
sheet3['A1'].font = font
font1 = Font(name='宋体', size='16', bold=True)
row4 = sheet3[4]
for ri in row4:
    ri.font = font1
    ri.alignment = ali
# 冻结表头
sheet3.topLeftCell = 'A1'
sheet3.freeze_panes = 'A5'
# 插入图表
chart = BarChart()
data = Reference(worksheet=sheet3, min_row=5, max_row=13, min_col=2, max_col=13)
cata = Reference(sheet3, min_row=5, max_row=13, min_col=1)
chart.add_data(data)
chart.set_categories(cata)
sheet3.add_chart(chart, 'A16')
# 保存表
workbook3.save(r'/home/gu/桌面/Python/venv/学习/venv/自动化.xlsx')
